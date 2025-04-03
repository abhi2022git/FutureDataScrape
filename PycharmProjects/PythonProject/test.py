# -------------------------------
# Import Necessary Libraries
# -------------------------------
import os  # For file handling
import time  # For delays
from openpyxl import load_workbook  # For working with Excel files
from openpyxl.styles import Font  # For Excel styling
import pandas as pd  # For data manipulation and analysis
import matplotlib.pyplot as plt  # For data visualization
from selenium import webdriver  # For web scraping
from selenium.webdriver.common.by import By  # For locating web elements
from selenium.webdriver.support.ui import WebDriverWait  # For dynamic waits
from selenium.webdriver.support import expected_conditions as EC
import numpy as np  # For numerical operations
from selenium.webdriver.chrome.service import Service  # Chrome WebDriver Service
from webdriver_manager.chrome import ChromeDriverManager  # WebDriver Manager for Chrome

# -------------------------------
# Selenium: Extract Data from Webpage
# -------------------------------
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# Target website URL
url = "https://www.barchart.com/futures"

# Navigate to the target webpage
driver.get(url)

# Wait for the page to load completely using explicit wait
try:
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'block-content')]"))
    )
except Exception as e:
    print("Error: Could not load webpage.")
    driver.quit()
    raise e

# Extract the main title of the page
title_element = driver.find_element(By.CLASS_NAME, "js-main-title")
title_text = title_element.text.strip()
print("Page Title:", title_text)

# Extract the raw data text from the specified element
raw_text = element.text.strip()

# -------------------------------
# Data Processing: Convert Raw Text to a DataFrame
# -------------------------------
lines = raw_text.split("\n")
header = lines[:8]
columns = header[:7]
data_lines = lines[8:]
rows = [data_lines[i:i+7] for i in range(0, len(data_lines), 7)]
df = pd.DataFrame(rows, columns=columns)
print("Formatted DataFrame:")
print(df.head(20))

# -------------------------------
# Save Processed Data to an Excel File with Mean Column
# -------------------------------
output_file = "extracted_data_with_mean.xlsx"

if os.path.exists(output_file):
    os.remove(output_file)

# Add a 'Mean' column
def convert_price(value):
    if '-' in str(value):
        main, fraction = value.split('-')
        return float(main) + float(fraction) / 8
    try:
        return float(value)
    except ValueError:
        return np.nan  # Return NaN for invalid values

df['High'] = pd.to_numeric(df['High'].apply(convert_price), errors='coerce')
df['Low'] = pd.to_numeric(df['Low'].apply(convert_price), errors='coerce')
df['Last'] = pd.to_numeric(df['Last'].apply(convert_price), errors='coerce')
df['Change'] = pd.to_numeric(df['Change'], errors='coerce')
df['Volume'] = pd.to_numeric(df['Volume'].str.replace(',', ''), errors='coerce')
df.dropna(subset=['High', 'Low', 'Change'], inplace=True)
df = df[df['High'] >= df['Low']]
df['Mean'] = (df['High'] + df['Low']) / 2

# Save the updated DataFrame to an Excel file, including 'Mean'
df.to_excel(output_file, sheet_name="Processed Data", index=False)
print(f"Data successfully saved to {output_file} with the 'Mean' column.")

# -------------------------------
# Add Page Title and Largest Change Analysis
# -------------------------------
wb = load_workbook(output_file)
sheet = wb["Processed Data"]

sheet.insert_rows(1)
sheet.cell(row=1, column=1).value = "Page Title:"
sheet.cell(row=1, column=2).value = title_text
header_font = Font(bold=True)
sheet.cell(row=1, column=1).font = header_font
sheet.cell(row=1, column=2).font = header_font

if 'Last' in df.columns and 'Change' in df.columns:
    max_change_idx = df['Change'].idxmax()
    max_change_row = df.loc[max_change_idx, ['Contract Name', 'Last']]
    last_row = sheet.max_row
    sheet.cell(row=last_row + 2, column=1).value = "Row with the Largest Change:"
    sheet.cell(row=last_row + 3, column=1).value = max_change_row['Contract Name']
    sheet.cell(row=last_row + 3, column=2).value = max_change_row['Last']
    sheet.cell(row=last_row + 2, column=1).font = header_font

wb.save(output_file)
print("Page Title and Row with the Largest Change successfully added to Excel file.")

# -------------------------------
# Visualization: Plot 'High', 'Low', and 'Mean'
# -------------------------------
plt.figure(figsize=(10, 6))
plt.plot(df.index, df['High'], marker='o', label='High')
plt.plot(df.index, df['Low'], marker='o', label='Low')
plt.plot(df.index, df['Mean'], marker='o', label='Mean')
plt.xlabel('Row Number (Index)')
plt.ylabel('Price')
plt.title('High, Low, and Mean Prices')
plt.legend()
plt.grid(True)
plt.tight_layout()
plt.show()